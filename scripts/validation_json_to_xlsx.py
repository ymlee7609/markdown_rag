#!/usr/bin/env python3
"""검증 데이터셋 JSON을 엑셀(.xlsx)로 변환한다.

사용법:
    python3 scripts/validation_json_to_xlsx.py [입력.json] [출력.xlsx]

기본값: scripts/validation_dataset_manual.json -> scripts/validation_dataset_manual.xlsx

시트 구성:
    테스트케이스 - 케이스 1건 = 1행 (검토/리뷰용 메인 시트)
    근거         - 정답 케이스의 근거 파일/행/원문
    부재근거     - 부정 케이스의 grep 재현 명령과 매칭 수
    커버리지     - 벤더 x 유형 매트릭스
    메타         - 데이터셋 버전, 합격 기준, 유형 정의
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
POS_FILL = PatternFill("solid", fgColor="E8F1DE")  # 정답 케이스
NEG_FILL = PatternFill("solid", fgColor="FCE9E9")  # 부정 케이스
TITLE_FONT = Font(bold=True, size=10)

# 유형 코드 -> 한글 표기
CATEGORY_KO = {
    "spec_value": "제원·기본값",
    "command_syntax": "명령어 문법",
    "diagnostics": "진단·상태확인",
    "negative": "부정 케이스",
}
FLAVOR_KO = {
    "missing_doc_type": "문서종류 부재",
    "substring_false_match": "부분문자열 오검색",
    "cross_vendor_syntax": "교차벤더 문법",
    "sibling_model_leak": "형제모델 전이",
    "out_of_corpus_spec": "코퍼스 전체 부재",
}


def join(value) -> str:
    """리스트/None을 셀에 넣을 문자열로 정규화한다."""
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return str(value)


def style_sheet(ws, widths, wrap_cols=(), freeze="A2"):
    """헤더 서식, 열 너비, 자동필터, 틀 고정을 일괄 적용한다."""
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            wrap = cell.column in wrap_cols
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = freeze


def sheet_cases(wb, data):
    ws = wb.create_sheet("테스트케이스")
    ws.append([
        "ID", "벤더", "모델", "출처", "유형", "극성", "부정유형", "질의",
        "기대 경로(포함)", "기대 키워드", "기대 동작", "기대 답변",
        "오답 예시", "답변 금지 문구", "인용 금지 경로", "검증 포인트",
    ])
    for case in data["test_cases"]:
        negative = case["polarity"] == "negative"
        ws.append([
            case["id"],
            case["vendor"],
            case.get("model", ""),
            "큐레이션" if case.get("source") == "curated" else "생성",
            CATEGORY_KO.get(case["category"], case["category"]),
            "정답" if not negative else "부정",
            FLAVOR_KO.get(case.get("negative_flavor", ""), case.get("negative_flavor", "")),
            case["query"],
            case.get("expected_path_contains", ""),
            join(case.get("expected_keywords")),
            case.get("expected_behavior", "answer"),
            case.get("expected_answer", ""),
            case.get("wrong_answer_example", ""),
            join(case.get("fail_if_answer_contains")),
            join(case.get("must_not_cite_path_contains")),
            case.get("notes", ""),
        ])
        fill = NEG_FILL if negative else POS_FILL
        for cell in ws[ws.max_row]:
            cell.fill = fill
    style_sheet(
        ws,
        widths=[18, 13, 13, 9, 14, 7, 17, 52, 42, 34, 13, 58, 44, 30, 20, 70],
        wrap_cols=(8, 9, 10, 12, 13, 14, 15, 16),
    )
    return ws


def sheet_evidence(wb, data):
    ws = wb.create_sheet("근거")
    ws.append(["케이스 ID", "벤더", "유형", "근거 파일", "행", "끝행", "인용유형", "원문 발췌"])
    for case in data["test_cases"]:
        for ev in case.get("evidence", []):
            ws.append([
                case["id"],
                case["vendor"],
                CATEGORY_KO.get(case["category"], case["category"]),
                ev["file"],
                ev["line"],
                ev.get("line_end", ""),
                "축자" if ev.get("quote_type", "verbatim") == "verbatim" else "요약",
                ev["text"],
            ])
    style_sheet(ws, widths=[18, 13, 14, 58, 7, 7, 10, 100], wrap_cols=(4, 8))
    return ws


def sheet_absence(wb, data):
    ws = wb.create_sheet("부재근거")
    ws.append(["케이스 ID", "벤더", "부정유형", "재현 grep 명령", "매칭 수", "비고"])
    for case in data["test_cases"]:
        absence = case.get("absence_evidence")
        if not absence:
            continue
        note = absence.get("note", "")
        owner = absence.get("correct_owner") or absence.get("correct_syntax_ref")
        if owner:
            note = (note + " / 실제 소유: " + json.dumps(owner, ensure_ascii=False)).strip(" /")
        ws.append([
            case["id"],
            case["vendor"],
            FLAVOR_KO.get(case.get("negative_flavor", ""), case.get("negative_flavor", "")),
            absence.get("command", ""),
            absence.get("match_count", ""),
            note,
        ])
    style_sheet(ws, widths=[18, 13, 17, 70, 9, 80], wrap_cols=(4, 6))
    return ws


def sheet_coverage(wb, data):
    ws = wb.create_sheet("커버리지")
    cats = ["spec_value", "command_syntax", "diagnostics", "negative"]
    ws.append(["벤더"] + [CATEGORY_KO[c] for c in cats] + ["합계"])
    grid = {}
    for case in data["test_cases"]:
        v = grid.setdefault(case["vendor"], {})
        v[case["category"]] = v.get(case["category"], 0) + 1
    for vendor in data["vendors"]:
        row = grid.get(vendor, {})
        ws.append([vendor] + [row.get(c, 0) for c in cats] + [sum(row.values())])
    ws.append(["합계"] + [sum(1 for c in data["test_cases"] if c["category"] == cat) for cat in cats]
              + [len(data["test_cases"])])
    for cell in ws[ws.max_row]:
        cell.font = TITLE_FONT
    style_sheet(ws, widths=[15, 20, 20, 20, 20, 8])
    return ws


def sheet_meta(wb, data):
    ws = wb.create_sheet("메타")
    ws.append(["항목", "값"])
    rows = [
        ("version", data.get("version", "")),
        ("description", data.get("description", "")),
        ("corpus_root", data.get("corpus_root", "")),
        ("벤더 수", len(data.get("vendors", []))),
        ("케이스 수", len(data["test_cases"])),
        ("정답 케이스", sum(1 for c in data["test_cases"] if c["polarity"] == "positive")),
        ("부정 케이스", sum(1 for c in data["test_cases"] if c["polarity"] == "negative")),
    ]
    for key, value in data.get("pass_criteria", {}).items():
        rows.append((f"pass_criteria.{key}", value))
    for key, value in data.get("category_definitions", {}).items():
        rows.append((f"유형정의.{CATEGORY_KO.get(key, key)}", value))
    for key, value in data.get("negative_flavors", {}).items():
        rows.append((f"부정유형.{FLAVOR_KO.get(key, key)}", value))
    for key, value in rows:
        ws.append([key, str(value)])
    style_sheet(ws, widths=[34, 120], wrap_cols=(2,))
    return ws


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else root / "scripts/validation_dataset_manual.json"
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else src.with_suffix(".xlsx")

    data = json.loads(src.read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)
    sheet_cases(wb, data)
    sheet_evidence(wb, data)
    sheet_absence(wb, data)
    sheet_coverage(wb, data)
    sheet_meta(wb, data)
    wb.save(dst)

    print(f"생성: {dst}")
    for ws in wb.worksheets:
        print(f"  - {ws.title}: {ws.max_row - 1}행 x {ws.max_column}열")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
